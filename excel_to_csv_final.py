import openpyxl
import csv
import os
import re
import shutil

""" 
El primer bloque es para setear. La idea es crear dos directorios
 """
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
# Crear directorios si no existen
path = os.getcwd()

#creanmos los direcotrios Excel y CSV
os.makedirs('csv_files', exist_ok=True)
#creamos un variable para usar con con el path
new_path = path + '\\' + 'csv_files'
#---FUNCIONES---#
for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):
        continue
    workbook = openpyxl.load_workbook(excel_file)
    #loop para pasar las sheets en el workbook
    for sheets in workbook.sheetnames: #sintaxis diferente que el libro
        wb_name = re.sub('.xlsx', '', excel_file)  # Eliminar la extensión .xls
        csv_name = wb_name + '_' + sheets + '.csv'  # Aca renonbramos coml csv
        #Abrimos el csv y creamos nuestro WRITE OBJECT
        cvs_file=open(csv_name,'w', newline='')
        csv_writer= csv.writer(cvs_file)
        sheet= workbook.active
        
        #Creamos una nueva lista por cada ROW en cada SHEET
        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell_data= sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_data)
            #Este es eel paso de la escritura ,  aca escribe el row de una
            csv_writer.writerow(row_data)
        #Cerramos el csv
        #Movemos el cvsa del directorio a uno dedicado
        cvs_file.close()
        shutil.move(os.path.join(path, csv_name), os.path.join(new_path,csv_name))

print('Listo')
           



#---NOTAS---#
#Eecutar desde el cmd ,luego de instalacion de phyton , where phyton , where pip, where excel_to_csv.py 
#cambiar la direccion del script, en este ejemplo es mi carpeta personal
#Si hay problemas revisar el PATH en variables de entorno .
#Lo mismo en consifguracion,aplicaciones , opciones avanzadas , alias de ejecucion, desacativar los python.exe y phython3.exe
#Luego en CMD :pip install openpyxl 
#Ejecucion:python excel_to_csv.py


##Proxima modificacion:
#directorios dedicados ,  traer los excel de otras carpetas sin necesisda de pegarlas en la carpeta de este script

